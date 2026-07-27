"""
Scan the raw survey export sheets for question-code tokens used by the
design template (S1, S2a, S2b, S3, S4, S9, QA2, QA4, SZ, CO2, CX, CX2, SOA, SOI)
and report which sheet/column each one lives in, plus a sample of header text.

This is read-only reconnaissance to de-risk building the data conversion
script -- survey platform exports embed question codes as a prefix inside a
long free-text header, so we search rather than assume fixed positions.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

SOURCE = Path(__file__).resolve().parent.parent / "Data Source" / "P. Hansel BHT - Progress Report - V2 (16072026)-Checked RF.xlsx"
SHEETS = ["BHT", "Campaign Evaluation", "Young Generation"]
CODES = ["S1", "S2a", "S2b", "S3", "S4", "S9", "QA2", "QA4", "SZ", "CO2", "CX2", "CX", "SOA", "SOI"]

# Match "S3." / "S3 " / "S3)" etc at the start of the header text, longest-code-first
# (CX2 before CX) so we don't mis-tag a CX2 column as CX.
CODE_PATTERN = re.compile(
    r"^\s*(" + "|".join(re.escape(c) for c in sorted(CODES, key=len, reverse=True)) + r")[\s.\):]"
)


def header_rows(ws, n=2):
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=n, values_only=True), start=1):
        rows.append(row)
    return rows


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    report = {code: [] for code in CODES}

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = header_rows(ws, n=2)
        row1 = rows[0] if len(rows) > 0 else ()
        row2 = rows[1] if len(rows) > 1 else ()
        max_col = max(len(row1), len(row2))
        for col_idx in range(max_col):
            r1 = row1[col_idx] if col_idx < len(row1) else None
            r2 = row2[col_idx] if col_idx < len(row2) else None
            header_text = str(r2 if r2 not in (None, "") else r1 or "")
            m = CODE_PATTERN.match(header_text)
            if m:
                code = m.group(1)
                report[code].append({
                    "sheet": sheet_name,
                    "col_index_1based": col_idx + 1,
                    "row1": r1,
                    "row2": r2,
                })

    print(json.dumps(report, indent=2, ensure_ascii=False, default=str))

    unmatched = [c for c, hits in report.items() if not hits]
    if unmatched:
        print(f"\n# NOTE: no columns found for codes: {unmatched}", file=sys.stderr)


if __name__ == "__main__":
    main()
